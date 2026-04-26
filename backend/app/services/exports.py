from __future__ import annotations

from collections import Counter
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from sqlalchemy.orm import Session

from ..models import Ticket


EXPORT_COLUMNS = [
    ("Ticket ID", "vendor_ticket_id"),
    ("Subject", "vendor_subject"),
    ("Vendor Status", "vendor_status"),
    ("Internal Status", "internal_status"),
    ("Internal Owner", "internal_owner"),
    ("Vendor Issue Category", "vendor_issue_category"),
    ("Issue Category", "issue_category"),
    ("Root Cause", "root_cause"),
    ("From", "vendor_from_name"),
    ("From Email", "vendor_from_email"),
    ("Help Topic", "vendor_help_topic"),
    ("Priority", "vendor_priority"),
    ("Created Date", "vendor_created_date"),
    ("Closed Date", "vendor_closed_date"),
    ("Comments", "comments"),
]

BRAND_PRIMARY = "002A3A"
BRAND_HEADER_BG = "EAF2F4"
LOGO_PATH = Path(__file__).resolve().parents[3] / "frontend" / "public" / "blatchford-mark.jpeg"


def is_open_ticket(ticket: Ticket) -> bool:
    return (ticket.vendor_status or "").strip().lower() != "closed"


def is_workshop_required(ticket: Ticket) -> bool:
    return (ticket.issue_category or "").strip().lower() == "workshop required"


def format_cell_value(value: object) -> object:
    if hasattr(value, "strftime"):
        return value
    return value


def add_sheet_branding(sheet, title: str) -> None:
    last_column = sheet.cell(row=1, column=len(EXPORT_COLUMNS)).column_letter
    sheet.row_dimensions[1].height = 48
    sheet.row_dimensions[2].height = 48
    sheet.merge_cells(f"C1:{last_column}2")

    title_cell = sheet["C1"]
    title_cell.value = f"Blatchford PLM Ticket Manager\n{title}"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)

    accent_cell = sheet["A1"]
    accent_cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)
    sheet["B1"].fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 16

    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH))
        logo.width = 76
        logo.height = 76
        logo.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=1,
                colOff=pixels_to_EMU(14),
                row=0,
                rowOff=pixels_to_EMU(26),
            ),
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(logo.width),
                cy=pixels_to_EMU(logo.height),
            ),
        )
        sheet.add_image(logo)


def write_sheet(workbook: Workbook, title: str, tickets: list[Ticket]) -> None:
    sheet = workbook.create_sheet(title)
    add_sheet_branding(sheet, title)

    headers = [column_title for column_title, _ in EXPORT_COLUMNS]
    sheet.append([])
    sheet.append(headers)

    for cell in sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ticket in tickets:
        sheet.append([format_cell_value(getattr(ticket, field_name)) for _, field_name in EXPORT_COLUMNS])

    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{sheet.cell(row=4, column=len(EXPORT_COLUMNS)).column_letter}{sheet.max_row}"

    for index, column_cells in enumerate(sheet.columns, start=1):
        if index == 1:
            continue
        if index == 2:
            continue
        values = [str(cell.value or "") for cell in column_cells]
        max_length = min(max(len(value) for value in values) + 2, 40)
        sheet.column_dimensions[get_column_letter(index)].width = max_length

    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
        for cell in row:
            if hasattr(cell.value, "strftime"):
                cell.number_format = "DD/MM/YYYY HH:MM"


def write_explanation_block(
    sheet,
    start_row: int,
    start_col: int,
    summary: str,
    source: str,
    message: str,
) -> int:
    entries = [
        ("Summary", summary),
        ("Source", source),
        ("Message", message),
    ]
    current_row = start_row
    for label, text in entries:
        label_cell = sheet.cell(row=current_row, column=start_col, value=f"{label}:")
        label_cell.font = Font(bold=True)
        sheet.cell(row=current_row, column=start_col + 1, value=text)
        current_row += 1
    return current_row + 1


def write_metric_card(sheet, row: int, column: int, title: str, value: object) -> None:
    title_cell = sheet.cell(row=row, column=column, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_PRIMARY)
    title_cell.alignment = Alignment(horizontal="center")

    value_cell = sheet.cell(row=row + 1, column=column, value=value)
    value_cell.font = Font(bold=True, size=16)
    value_cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_HEADER_BG)
    value_cell.alignment = Alignment(horizontal="center")


def write_summary_table(sheet, start_row: int, start_col: int, title: str, headers: list[str], rows: list[tuple[object, ...]]) -> tuple[int, int]:
    title_cell = sheet.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=12)

    header_row = start_row + 1
    for offset, header in enumerate(headers):
        cell = sheet.cell(row=header_row, column=start_col + offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=BRAND_HEADER_BG)

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            sheet.cell(row=header_row + row_offset, column=start_col + col_offset, value=value)

    return header_row, len(rows)


def add_pie_chart(sheet, anchor_cell: str, start_col: int, header_row: int, data_rows: int, title: str) -> None:
    if data_rows < 1:
        return

    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.height = 7
    chart.width = 10

    labels = Reference(sheet, min_col=start_col, min_row=header_row + 1, max_row=header_row + data_rows)
    data = Reference(sheet, min_col=start_col + 1, min_row=header_row, max_row=header_row + data_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    sheet.add_chart(chart, anchor_cell)


def add_bar_chart(
    sheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
    title: str,
    x_axis_title: str,
    y_axis_title: str,
) -> None:
    if data_rows < 1:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.height = 7
    chart.width = 10
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    labels = Reference(sheet, min_col=start_col, min_row=header_row + 1, max_row=header_row + data_rows)
    data = Reference(sheet, min_col=start_col + 1, min_row=header_row, max_row=header_row + data_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    sheet.add_chart(chart, anchor_cell)


def write_summary_sheet(workbook: Workbook, tickets: list[Ticket]) -> None:
    sheet = workbook.create_sheet("Summary & Reporting")
    add_sheet_branding(sheet, "Summary & Reporting")
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False

    total_tickets = len(tickets)
    open_tickets = sum(1 for ticket in tickets if is_open_ticket(ticket))
    workshop_tickets = sum(1 for ticket in tickets if is_workshop_required(ticket))
    enriched_tickets = sum(1 for ticket in tickets if ticket.issue_category or ticket.internal_status or ticket.comments)

    write_metric_card(sheet, 4, 1, "Total Tickets", total_tickets)
    write_metric_card(sheet, 4, 3, "Open Tickets", open_tickets)
    write_metric_card(sheet, 4, 5, "Workshop Required", workshop_tickets)
    write_metric_card(sheet, 4, 7, "Enriched Tickets", enriched_tickets)

    write_explanation_block(
        sheet,
        start_row=7,
        start_col=1,
        summary="Overview of current ticket distribution and enrichment coverage.",
        source="All records currently stored in the PLM Ticket Manager database.",
        message="Use this sheet as the management summary tab before moving into detailed ticket tabs.",
    )

    vendor_status_counts = Counter((ticket.vendor_status or "Unspecified").strip() or "Unspecified" for ticket in tickets)
    issue_category_counts = Counter((ticket.issue_category or "Unspecified").strip() or "Unspecified" for ticket in tickets)
    owner_counts = Counter((ticket.internal_owner or "Unassigned").strip() or "Unassigned" for ticket in tickets)

    top_issue_rows = sorted(issue_category_counts.items(), key=lambda item: (-item[1], item[0]))
    top_owner_rows = sorted(owner_counts.items(), key=lambda item: (-item[1], item[0]))[:8]
    status_rows = sorted(vendor_status_counts.items(), key=lambda item: (-item[1], item[0]))

    header_row, data_rows = write_summary_table(
        sheet,
        start_row=12,
        start_col=8,
        title="Vendor Status Breakdown",
        headers=["Status", "Count"],
        rows=[(status, count) for status, count in status_rows],
    )
    add_pie_chart(sheet, "A12", 8, header_row, data_rows, "Vendor Status Distribution")
    write_explanation_block(
        sheet,
        start_row=21,
        start_col=1,
        summary="Shows how the vendor currently classifies ticket state.",
        source="`vendor_status` values from imported supplier records.",
        message="Use this to monitor open vs closed workload and watch for tickets waiting on external action.",
    )

    header_row, data_rows = write_summary_table(
        sheet,
        start_row=30,
        start_col=8,
        title="Internal Issue Categories",
        headers=["Issue Category", "Count"],
        rows=[(category, count) for category, count in top_issue_rows],
    )
    add_bar_chart(sheet, "A30", 8, header_row, data_rows, "Internal Issue Categories", "Issue Category", "Tickets")
    write_explanation_block(
        sheet,
        start_row=39,
        start_col=1,
        summary="Summarises how tickets are being classified internally.",
        source="`issue_category` values entered by Blatchford users.",
        message="This supports trend reporting and highlights workshop demand through the `Workshop Required` category.",
    )

    header_row, data_rows = write_summary_table(
        sheet,
        start_row=48,
        start_col=8,
        title="Internal Owner Allocation",
        headers=["Internal Owner", "Count"],
        rows=[(owner, count) for owner, count in top_owner_rows],
    )
    add_bar_chart(sheet, "A48", 8, header_row, data_rows, "Internal Owner Allocation", "Internal Owner", "Tickets")
    write_explanation_block(
        sheet,
        start_row=57,
        start_col=1,
        summary="Shows which internal owners are attached to the most tickets.",
        source="`internal_owner` values entered in the ticket detail form.",
        message="Use this view to balance workload and identify records still waiting for internal ownership.",
    )

    for column in range(1, 8):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    for column in range(8, 10):
        sheet.column_dimensions[get_column_letter(column)].width = 20


def build_excel_report(db: Session) -> bytes:
    tickets = db.query(Ticket).order_by(Ticket.vendor_created_date.desc()).all()

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(workbook, "All Tickets", tickets)
    write_sheet(workbook, "Open Tickets", [ticket for ticket in tickets if is_open_ticket(ticket)])
    write_sheet(workbook, "Workshop Required", [ticket for ticket in tickets if is_workshop_required(ticket)])
    write_summary_sheet(workbook, tickets)
    workbook.active = 0

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()
